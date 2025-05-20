import os
import json
import openpyxl
from PyQt5.QtWidgets import QTableWidgetItem
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtCore import Qt
from ..utils.color_utils import apply_tint
import re


class ExcelService:
    def __init__(self):
        self.accent_colors = {1: '3B4E87'}  # accent1 파랑 계열 RGB를 하드코딩

    def load_excel(self, path, table_widget):
        """엑셀 파일을 로드하여 테이블 위젯에 표시합니다."""
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows, cols = ws.max_row, ws.max_column
        
        if table_widget is not None:
            table_widget.blockSignals(True)
            table_widget.clear()
            table_widget.setRowCount(rows)
            table_widget.setColumnCount(cols)

            # 헤더 정보 추출
            header_row, header_map = self._extract_header_info(ws)
            
            # 품목 정보 추출
            items = self._extract_items(ws, header_row, header_map)

            # 셀 스타일 적용
            self._apply_cell_styles(ws, table_widget)

            # 병합 셀 처리
            self._handle_merged_cells(ws, table_widget)

            # 열 너비/행 높이 설정
            self._set_dimensions(ws, table_widget)

            table_widget.blockSignals(False)
        else:
            # UI 없이 데이터만 추출
            header_row, header_map = self._extract_header_info(ws)
            items = self._extract_items(ws, header_row, header_map)

        # JSON 파일 저장
        json_path = os.path.splitext(path)[0] + ".json"
        self._save_to_json(json_path, ws, items)
        return json_path, items

    def normalize_header(self, text):
        return re.sub(r'[^a-z0-9가-힣]', '', text.lower()) if text else ''

    def _extract_header_info(self, ws, max_header_row=40):
        header_row = None
        header_map = {}
        unit_krw_candidates = [
            "unitkrw", "unitcost", "unitprice", "단가", "단가원", "단가$", "unit", "unitkrw", "unit$"
        ]
        quantity_candidates = ["qty", "quantity", "수량"]
        amount_candidates = ["amount", "totalamount", "금액", "합계"]
        description_candidates = ["description", "item", "품목", "상세내역"]
        unit_candidates = ["unit", "단위"]
        remark_candidates = ["remark", "비고", "note"]
        for r in range(1, min(ws.max_row, max_header_row) + 1):
            row_values = [
                str(ws.cell(row=r, column=c).value).strip() if ws.cell(row=r, column=c).value is not None else ""
                for c in range(1, ws.max_column + 1)
            ]
            norm_row = [self.normalize_header(cell) for cell in row_values]
            for idx, norm_cell in enumerate(norm_row):
                if norm_cell in description_candidates:
                    header_map["description"] = idx
                if norm_cell in unit_krw_candidates:
                    header_map["unit_krw"] = idx
                if norm_cell in quantity_candidates:
                    header_map["quantity"] = idx
                if norm_cell in amount_candidates:
                    header_map["amount"] = idx
                if norm_cell in unit_candidates:
                    header_map["unit"] = idx
                if norm_cell in remark_candidates:
                    header_map["remark"] = idx
            if len(header_map) >= 2:
                header_row = r
                break
        if not header_row or not header_map:
            print(f"[헤더 인식 실패] row_values: {row_values}")
            print(f"[헤더 인식 실패] header_map: {header_map}")
        return header_row, header_map

    def clean_number(self, val):
        try:
            if val is None or val == '':
                return None
            fval = float(val)
            if fval.is_integer():
                return int(fval)
            return round(fval, 2)
        except Exception:
            return val

    def _extract_items(self, ws, header_row, header_map):
        items = []
        if not header_row or not header_map:
            return items
        for r in range(header_row + 1, ws.max_row + 1):
            row_values = [
                str(ws.cell(row=r, column=c).value).strip()
                if ws.cell(row=r, column=c).value is not None else ""
                for c in range(1, ws.max_column + 1)
            ]
            desc_idx = header_map.get("description")
            if desc_idx is None or not row_values[desc_idx]:
                continue
            # summary 키워드 스킵
            if any(k in row_values[desc_idx].lower() for k in ["합계", "총액", "vat", "참고", "소계", "tax", "total", "sum"]):
                continue
            item = {
                "description": row_values[desc_idx],
                "unit_krw": self.clean_number(row_values[header_map["unit_krw"]]) if "unit_krw" in header_map else None,
                "quantity": self.clean_number(row_values[header_map["quantity"]]) if "quantity" in header_map else None,
                "unit": row_values[header_map["unit"]] if "unit" in header_map else None,
                "amount": self.clean_number(row_values[header_map["amount"]]) if "amount" in header_map else None,
                "remark": row_values[header_map["remark"]] if "remark" in header_map else None
            }
            items.append(item)
        return items

    def _apply_cell_styles(self, ws, table_widget):
        """셀 스타일을 적용합니다."""
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                text = cell.value or ""
                item = QTableWidgetItem(str(text))

                # 폰트
                f = cell.font
                point_size = int(f.sz) if f.sz is not None else -1
                qf = QFont(f.name, point_size)
                qf.setBold(f.b)
                qf.setItalic(f.i)
                item.setFont(qf)

                # 배경색
                color = self._get_cell_color(cell)
                if (color and 
                    (color.red(), color.green(), color.blue()) != (255, 255, 255)):
                    item.setBackground(color)

                # 정렬
                align = cell.alignment
                qt_align = 0
                if align.horizontal == 'center':
                    qt_align |= Qt.AlignHCenter
                elif align.horizontal == 'right':
                    qt_align |= Qt.AlignRight
                else:
                    qt_align |= Qt.AlignLeft
                if align.vertical == 'center':
                    qt_align |= Qt.AlignVCenter
                elif align.vertical == 'bottom':
                    qt_align |= Qt.AlignBottom
                else:
                    qt_align |= Qt.AlignTop
                item.setTextAlignment(qt_align)

                # 테두리
                b = cell.border
                border_info = {
                    "top": bool(b.top and b.top.style),
                    "bottom": bool(b.bottom and b.bottom.style),
                    "left": bool(b.left and b.left.style),
                    "right": bool(b.right and b.right.style),
                }
                item.setData(Qt.UserRole, border_info)
                table_widget.setItem(r-1, c-1, item)

    def _get_cell_color(self, cell):
        """셀의 배경색을 가져옵니다."""
        try:
            fill = cell.fill
            fg = fill.fgColor if hasattr(fill, 'fgColor') else None
            
            if (fill.patternType in ('solid', 'gray125', 'darkGrid', 'lightGrid')
                    and fg):
                # 1) Theme 컬러 + 틴트
                if fg.type == 'theme':
                    tint = getattr(fg, 'tint', 0.0)
                    hex_rgb = self.accent_colors[1]
                    return apply_tint(hex_rgb, tint)
                # 2) RGB 컬러
                elif fg.type == 'rgb' and fg.rgb:
                    rgb = fg.rgb[2:] if fg.rgb.startswith('FF') else fg.rgb
                    return QColor(
                        int(rgb[0:2], 16),
                        int(rgb[2:4], 16),
                        int(rgb[4:6], 16)
                    )
                # 3) Indexed 컬러
                elif fg.type == 'indexed' and fg.indexed is not None:
                    from openpyxl.styles.colors import COLOR_INDEX
                    idx = fg.indexed
                    if 0 <= idx < len(COLOR_INDEX):
                        hexcol = COLOR_INDEX[idx][2:]
                        return QColor(
                            int(hexcol[0:2], 16),
                            int(hexcol[2:4], 16),
                            int(hexcol[4:6], 16)
                        )
                # 4) Gradient Fill
                elif hasattr(fill, 'gradientType') and fill.gradientType:
                    stops = getattr(fill, 'stop', None)
                    if (stops and hasattr(stops[0], 'color')
                            and hasattr(stops[0].color, 'rgb')):
                        rgb = (stops[0].color.rgb[2:]
                               if stops[0].color.rgb.startswith('FF')
                               else stops[0].color.rgb)
                        return QColor(
                            int(rgb[0:2], 16),
                            int(rgb[2:4], 16),
                            int(rgb[4:6], 16)
                        )
        except Exception:
            pass
        return None

    def _handle_merged_cells(self, ws, table_widget):
        """병합된 셀을 처리합니다."""
        for merged in ws.merged_cells.ranges:
            r0, c0 = merged.min_row-1, merged.min_col-1
            rs = merged.max_row - merged.min_row + 1
            cs = merged.max_col - merged.min_col + 1
            table_widget.setSpan(r0, c0, rs, cs)

    def _set_dimensions(self, ws, table_widget):
        """열 너비와 행 높이를 설정합니다."""
        # 열 너비
        for idx, col_dim in ws.column_dimensions.items():
            col = openpyxl.utils.column_index_from_string(idx) - 1
            if col < ws.max_column and col_dim.width:
                table_widget.setColumnWidth(col, int(col_dim.width * 7))
        # 행 높이
        for r, row_dim in ws.row_dimensions.items():
            if row_dim.height is not None and r-1 < ws.max_row:
                table_widget.setRowHeight(r-1, int(row_dim.height * 1.2))

    def _save_to_json(self, json_path, ws, items):
        """데이터를 JSON 파일로 저장합니다."""
        data_dict = {
            "meta": {
                "file_name": os.path.basename(json_path)
            },
            "items": items,
            "header": {},
            "summary": {}
        }

        # 헤더 정보 추출
        header_labels = [
            "DATE", "QUOTATION #", "Payment date", "SHIP TO",
            "발급일", "공급자", "등록번호", "상호", "대표이사", "사업자 주소"
        ]
        for r in range(1, 16):
            for c in range(1, ws.max_column):
                raw = ws.cell(row=r, column=c).value
                if raw is None:
                    continue
                key = str(raw).strip().rstrip(":")
                if key in header_labels:
                    val = ws.cell(row=r, column=c+1).value \
                        or ws.cell(row=r+1, column=c).value
                    data_dict["header"][key] = val

        # 요약 정보 추출
        summary_labels = {
            "subtotal": ["Sub total", "소계"],
            "tax_rate": ["Tax rate", "세율"],
            "tax_due": ["Tax due", "세금", "Tax due"],
            "other": ["Other", "기타"],
            "total_due": ["TOTAL Due", "총액", "합계", "TOTAL"]
        }
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column):
                raw = ws.cell(row=r, column=c).value
                if raw is None:
                    continue
                txt = str(raw).strip()
                for fld, keys in summary_labels.items():
                    if any(k in txt for k in keys):
                        data_dict["summary"][fld] = (
                            ws.cell(row=r, column=c+1).value
                        )

        # 상단 고정 레이블 추출
        for r in range(3, 10):
            raw_key = ws.cell(row=r, column=4).value
            if raw_key:
                key = str(raw_key).strip().rstrip(':')
                data_dict["header"][key] = ws.cell(row=r, column=5).value

        # JSON 파일 저장
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data_dict, f, ensure_ascii=False, indent=2, default=str) 