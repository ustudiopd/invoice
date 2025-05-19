import os
import json
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import coordinate_to_tuple
from copy import copy
from openpyxl import Workbook

TEMPLATE_PATH = 'new_invoice_form.xlsx'
JSON_DIR = './2025년 견적서_주식회사/json'
OUTPUT_DIR = './converted_invoices'

os.makedirs(OUTPUT_DIR, exist_ok=True)

# 건너뛸 키워드 목록
SKIP_KEYWORDS = [
    'Thank You', 'U-STUDIO', 'support@ustudio.co.kr',
    'OTHER COMMENTS', 'Tax rate', 'Tax due', 'Other',
    'TOTAL Due', '합계', 'Sub Total', 'Delivery',
    '추가 사항', '기타 사항', 'Category', 'Etc'
]

SUMMARY_KEYWORDS = ['합계', '총액', 'vat', '참고', 'sum', 'total']

CATEGORY_FILL = PatternFill(
    start_color='E6F3FF',
    end_color='E6F3FF',
    fill_type='solid'
)
NORMAL_FILL = PatternFill(
    start_color='FFFFFF',
    end_color='FFFFFF',
    fill_type='solid'
)

def is_skip_row(desc):
    if not desc:
        return True
    desc = str(desc).strip()
    return any(kw in desc for kw in SKIP_KEYWORDS)

def is_summary_row(desc):
    if not desc:
        return False
    desc = str(desc).strip().lower()
    return any(kw in desc for kw in SUMMARY_KEYWORDS)

def write_amount(cell, value):
    if value is None:
        cell.value = None
        return
    try:
        if isinstance(value, str):
            value = value.replace(',', '')
        cell.value = float(value)
    except (ValueError, TypeError):
        cell.value = None

def set_cell_value(ws, cell_coordinate, value):
    if value is None:
        value = ''
    row, col = coordinate_to_tuple(cell_coordinate)
    for merged_range in ws.merged_cells.ranges:
        if (row, col) in merged_range.cells:
            master_row, master_col = merged_range.min_row, merged_range.min_col
            if (row, col) != (master_row, master_col):
                return
            try:
                ws.cell(row=master_row, column=master_col).value = value
                return
            except Exception as e:
                print(f"ERROR: 마스터 셀 {ws.cell(row=master_row, column=master_col).coordinate}에 값 쓰기 실패: {e}")
                raise
    try:
        ws[cell_coordinate].value = value
    except Exception as e:
        print(f"ERROR: {cell_coordinate}에 값 쓰기 실패: {e}")
        raise

def copy_row_style(ws, src_row, dst_row, max_col=5):
    for col in range(1, max_col+1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)

def map_to_new_template(item):
    desc = (
        item.get('description') 
        or item.get('품명') 
        or item.get('item') 
        or item.get('item_name')
        or item.get('품목명')
        or item.get('name')
        or '—항목 미지정—'
    )
    return {
        'description': desc,
        'unit_krw': item.get('unit_krw') or item.get('단가') or item.get('unit_cost', ''),
        'qty': item.get('qty') or item.get('수량') or item.get('quantity', ''),
        'unit': item.get('unit', ''),
        'amount_krw': item.get('amount_krw') or item.get('금액') or item.get('total_amount', ''),
        'remark': item.get('비고') or item.get('remark', ''),
    }

def flatten_items(raw_items):
    # 카테고리 구조인지, 단일 리스트인지 자동 판별
    if (
        isinstance(raw_items, list)
        and raw_items
        and isinstance(raw_items[0], dict)
        and 'category' in raw_items[0]
        and 'items' in raw_items[0]
    ):
        # 카테고리 구조
        flat = []
        for cat in raw_items:
            category_name = cat.get('category', '')
            if category_name:
                # 카테고리명만 가진 dict 추가
                flat.append({'description': category_name})
            for item in cat.get('items', []):
                flat.append(item)
        return flat
    # 단일 리스트 구조
    return raw_items

def fill_invoice_from_json(json_path, template_path, output_path):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row >= 12 and merged_range.max_row <= 44 and
            merged_range.min_col >= 1 and merged_range.max_col <= 6):
            ws.unmerge_cells(str(merged_range))
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    header = data.get('header', {})
    raw_items = data.get('items', [])
    # items 구조 자동 flatten
    flat_items = flatten_items(raw_items)
    mapped_items = [map_to_new_template(item) for item in flat_items]
    if 'date' in header:
        set_cell_value(ws, 'F4', header['date'])
    if 'quotation_no' in header:
        set_cell_value(ws, 'F5', header['quotation_no'])
    if 'payment_date' in header:
        set_cell_value(ws, 'F6', header['payment_date'])
    if 'ship_to' in header:
        set_cell_value(ws, 'F7', header['ship_to'])
    start_row = 12
    max_table_row = 44
    n_data = len(mapped_items)
    n_template = max_table_row - start_row + 1
    if n_data > n_template:
        for i in range(n_template, n_data):
            ws.insert_rows(max_table_row + i - n_template + 1)
            copy_row_style(ws, max_table_row, max_table_row + i - n_template + 1, max_col=6)
    data_rows = []
    summary_rows = []
    for mapped in mapped_items:
        if is_summary_row(mapped['description']):
            summary_rows.append(mapped)
        else:
            data_rows.append(mapped)
    for idx, mapped in enumerate(data_rows):
        row = start_row + idx
        # description만 있고 나머지 값이 없으면 카테고리 행으로 간주
        if mapped['description'] and not any([mapped['unit_krw'], mapped['qty'], mapped['unit'], mapped['amount_krw'], mapped['remark']]):
            ws.merge_cells(f'A{row}:F{row}')
            set_cell_value(ws, f'A{row}', mapped['description'])
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = CATEGORY_FILL
        else:
            set_cell_value(ws, f'A{row}', mapped['description'])
            set_cell_value(ws, f'B{row}', mapped['unit_krw'])
            set_cell_value(ws, f'C{row}', mapped['qty'])
            set_cell_value(ws, f'D{row}', mapped['unit'])
            set_cell_value(ws, f'E{row}', mapped['amount_krw'])
            set_cell_value(ws, f'F{row}', mapped['remark'])
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = NORMAL_FILL
    for idx, mapped in enumerate(summary_rows):
        row = start_row + len(data_rows) + idx
        set_cell_value(ws, f'A{row}', mapped['description'])
        set_cell_value(ws, f'E{row}', mapped['amount_krw'])
        for col in ['B', 'C', 'D', 'F']:
            set_cell_value(ws, f'{col}{row}', '')
    last_data_row = start_row + n_data
    if last_data_row <= max_table_row:
        ws.delete_rows(last_data_row, max_table_row - last_data_row + 1)
    wb.save(output_path)
    print(f"✅ 변환 완료: {output_path}")

def json_to_table_excel(json_path, output_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    items = data.get('items', [])
    wb = Workbook()
    ws = wb.active
    # 표 컬럼명
    columns = ['Description', 'Unit KRW', 'QTY', 'UNIT', 'Amount', 'Remark']
    ws.append(columns)
    def write_item_row(item):
        ws.append([
            item.get('description', ''),
            item.get('unit_krw', ''),
            item.get('qty', ''),
            item.get('unit', ''),
            item.get('amount_krw', ''),
            item.get('remark', ''),
        ])
    # 카테고리 구조
    if (
        isinstance(items, list)
        and items
        and isinstance(items[0], dict)
        and 'category' in items[0]
        and 'items' in items[0]
    ):
        for cat in items:
            category_name = cat.get('category', '')
            ws.append([f'[{category_name}]'])
            for item in cat.get('items', []):
                write_item_row(item)
    else:
        for item in items:
            write_item_row(item)
    wb.save(output_path)
    print(f"✅ 표 형태로 저장 완료: {output_path}")

def main():
    for fname in os.listdir(JSON_DIR):
        if not fname.endswith('.json'):
            continue
        json_path = os.path.join(JSON_DIR, fname)
        base, ext = os.path.splitext(fname)
        output_path = os.path.join(
            OUTPUT_DIR,
            f"{base}_table.xlsx"
        )
        try:
            json_to_table_excel(json_path, output_path)
        except Exception as e:
            print(f"  → 변환 실패: {e}")

if __name__ == "__main__":
    main()