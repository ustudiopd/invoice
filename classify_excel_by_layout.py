import os
import shutil
import openpyxl

# 분류할 폴더 경로
base_dir = './2025년 견적서_주식회사'
output_base = './2025년 견적서_주식회사/분류결과'

os.makedirs(output_base, exist_ok=True)

def detect_layout_type(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        # 1. 컬럼명(10~20행)으로 판별 (우선)
        for row in ws.iter_rows(min_row=10, max_row=20):
            colnames = [str(cell.value).upper() if cell.value else '' for cell in row]
            if 'DESCRIPTION' in colnames and 'UNIT KRW' in colnames:
                return '영문INVOICE'
            if 'ITEM' in colnames and 'UNIT COST (KRW)' in colnames:
                return 'Quotation'
            if '상세내역' in colnames and '수량' in colnames:
                return '한글견적서'
        # 2. 타이틀/상단 텍스트로 판별 (보조)
        for row in ws.iter_rows(min_row=1, max_row=6):
            for cell in row:
                val = str(cell.value).upper() if cell.value else ''
                if 'INVOICE' in val:
                    return '영문INVOICE'
                if 'QUOTATION' in val:
                    return 'Quotation'
                if '견적서' in val:
                    return '한글견적서'
        return '기타'
    except Exception as e:
        print(f"분류 실패: {excel_path} ({e})")
        return '기타'

for fname in os.listdir(base_dir):
    if not fname.lower().endswith(('.xlsx', '.xls')):
        continue
    fpath = os.path.join(base_dir, fname)
    layout_type = detect_layout_type(fpath)
    out_dir = os.path.join(output_base, layout_type)
    os.makedirs(out_dir, exist_ok=True)
    shutil.copy2(fpath, os.path.join(out_dir, fname))
    print(f"{fname} → {layout_type}")

print("✅ 분류 완료") 