import os
import json
import openpyxl

class InvoiceExtractor:
    """견적서 엑셀 파일을 분석하여 JSON으로 변환하는 클래스"""
    
    def __init__(self):
        self.templates = {
            '영문 INVOICE': {
                'header': {
                    'company': 'A1',
                    'address': 'A2',
                    'vat_number': 'A8',
                    'date': 'E3',
                    'quotation_no': 'E4',
                    'po': 'E5',
                    'ship_to': 'E6',
                },
                'items': {
                    'start_row': 12,
                    'columns': {
                        'description': 'A',
                        'unit_krw': 'B',
                        'qty': 'C',
                        'unit': 'D',
                        'amount_krw': 'E',
                        'remark': 'F',
                    }
                },
                'summary': {
                    'total_krw': 'E18',
                    'tax_rate': 'E19',
                    'tax_due': 'E20',
                    'total_due_krw': 'E22',
                }
            },
            'Quotation': {
                'header': {
                    'title': 'B6',
                    'client': 'G7',
                    'purpose': 'G8',
                    'event_date': 'G9',
                    'manager': 'G10',
                },
                'items': {
                    'start_row': 11,
                    'columns': {
                        'description': 'B',
                        'qty': 'C',
                        'unit': 'D',
                        'unit_krw': 'E',
                        'amount_krw': 'F',
                        'remark': 'G',
                    }
                },
                'summary': {
                    'sub_total': 'F16',
                    'grand_total': 'F20',
                    'agency_commission': 'F21',
                    'cut': 'F22',
                    'grand_total_final': 'F23',
                }
            },
            '한글 견적서': {
                'header': {
                    'title': 'E1',
                    'date': 'E4',
                    'company': 'G3',
                    'vat_number': 'G4',
                    'ceo': 'G5',
                    'address': 'G6',
                },
                'items': {
                    'start_row': 13,
                    'columns': {
                        'description': 'E',
                        'qty': 'R',
                        'unit_krw': 'V',
                        'amount_krw': 'W',
                        'remark': 'Y',
                    }
                },
                'summary': {
                    '합계': 'W45',
                    'VAT': 'W46',
                    '총액': 'W47',
                }
            },
            '카테고리 견적서': {
                'header': {
                    'company': 'B2',
                    'quotation_title': 'F2',
                    'client': 'G3',
                    'purpose': 'G4',
                    'event_date': 'G5',
                    'manager': 'G6',
                },
                'items': {
                    'start_row': 10,
                    'columns': {
                        'category': 'B',
                        'item': 'C',
                        'quantity': 'D',
                        'day': 'E',
                        'unit_cost': 'F',
                        'total_amount': 'G',
                        'remark': 'H',
                    }
                },
                'summary': {
                    'grand_total': 'F28',
                    'agency_commission': 'F29',
                    'grand_total_final': 'F31',
                }
            },
            '거래명세서': {
                'header': {
                    'title': 'O2',
                    'date': 'D3',
                    'company': 'G4',
                    'vat_number': 'G5',
                    'ceo': 'G6',
                    'address': 'G7',
                },
                'items': {
                    'start_row': 13,
                    'columns': {
                        '품명': 'D',
                        '규격': 'E',
                        '수량': 'N',
                        '단가': 'V',
                        '금액': 'W',
                        '비고': 'Y',
                    }
                },
                'summary': {
                    '합계': 'W45',
                    'VAT': 'W46',
                    '총액': 'W47',
                }
            }
        }
    
    def detect_template_type(self, ws):
        # 1. 타이틀 탐색 (D1, E1, F1, G1 등)
        for col in ['D', 'E', 'F', 'G']:
            cell = f'{col}1'
            value = ws[cell].value
            if value and (
                'INVOICE' in str(value).upper() or
                'QUOTATION' in str(value).upper()
            ):
                return '영문 INVOICE'
        # 1-2. Quotation 양식 (B6~G6에 Quotation)
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            cell = f'{col}6'
            if ws[cell].value and 'QUOTATION' in str(ws[cell].value).upper():
                return 'Quotation'
        # 2. 컬럼명 탐색 (10~20행)
        for row in ws.iter_rows(min_row=10, max_row=20):
            colnames = [str(cell.value).upper() if cell.value else '' for cell in row]
            if (
                'DESCRIPTION' in colnames and
                'UNIT KRW' in colnames and
                'QTY' in colnames
            ):
                return '영문 INVOICE'
            if (
                'ITEM' in colnames and
                'UNIT COST (KRW)' in colnames and
                'TOTAL AMOUNT (KRW)' in colnames
            ):
                return 'Quotation'
        # 3. 카테고리 견적서
        for row in ws.iter_rows(min_row=8, max_row=12):
            for cell in row:
                if cell.value and 'Category' in str(cell.value):
                    return '카테고리 견적서'
        # 4. 한글 견적서/거래명세서
        for row in ws.iter_rows(min_row=1, max_row=6):
            for cell in row:
                if cell.value and ('견적서' in str(cell.value)):
                    return '한글 견적서'
                if cell.value and ('거래명세서' in str(cell.value)):
                    return '거래명세서'
        return '기타'
    
    def extract_header(self, ws, template):
        header = {}
        
        # 1. 템플릿에 정의된 셀 위치에서 추출
        for key, cell in template['header'].items():
            if cell in ws:
                value = ws[cell].value
                if value:
                    # 수식인 경우 계산된 값 사용
                    if isinstance(value, str) and value.startswith('='):
                        try:
                            value = ws[cell].value
                        except Exception:
                            pass
                    header[key] = str(value)
        
        # 2. 템플릿에 없는 경우 자동 탐색
        if not header:
            # 회사명/주소 탐색 (A1~A10)
            for row in range(1, 11):
                cell = f'A{row}'
                if ws[cell].value:
                    value = str(ws[cell].value)
                    if '주식회사' in value:
                        header['company'] = value
                    elif '사업자등록번호' in value:
                        header['vat_number'] = value
                    elif '대표' in value:
                        header['ceo'] = value
                    elif '주소' in value:
                        header['address'] = value
            
            # 날짜/견적번호 탐색 (E1~E10)
            for row in range(1, 11):
                cell = f'E{row}'
                if ws[cell].value:
                    value = str(ws[cell].value)
                    if '날짜' in value or 'Date' in value:
                        header['date'] = value
                    elif '견적번호' in value or 'Quotation No' in value:
                        header['quotation_no'] = value
        
        return header
    
    def get_cell_value(self, ws, row, col):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            return cell.value
        # 병합 셀 처리 (셀 범위 직접 비교)
        for merged in ws.merged_cells.ranges:
            if (merged.min_row <= row <= merged.max_row) and (merged.min_col <= col <= merged.max_col):
                return ws.cell(row=merged.min_row, column=merged.min_col).value
        return None
    
    def is_hangul_invoice(self, ws):
        # 12행에 '상세내역', '수량', '단가', '금액' 등 한글 컬럼이 모두 있으면 True
        header_row = 12
        colnames = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                colnames.append(str(cell_value).strip().replace('\n', '').replace(' ', '').replace('\xa0', '').upper())
            else:
                colnames.append('')
        return (
            '상세내역' in colnames and '수량' in colnames and '단가' in colnames and '금액' in colnames
        )

    def clean_number(self, val):
        try:
            if val is None or val == '':
                return None
            fval = float(val)
            if fval.is_integer():
                return int(fval)
            return round(fval, 2)
        except Exception:
            return val

    def extract_items(self, ws, template):
        categories = []
        current_category = None
        current_category_total = None
        current_items = []
        # 1. 컬럼명 행 자동 탐색 (5~30행)
        start_row = None
        columns = {}
        # 컬럼명 후보군(유사 컬럼명 포함, 공백/특수문자/대소문자 무시)
        def normalize(text):
            import re
            return re.sub(r'[^\w]', '', str(text).replace(' ', '').lower())
        desc_candidates = [normalize(x) for x in ['상세내역', '세부내역', '품명', '항목명', 'item', 'description', 'name']]
        qty_candidates = [normalize(x) for x in ['수량', 'qty', 'quantity']]
        unit_candidates = [normalize(x) for x in ['일수', 'unit', 'day', 'days', '단위']]
        unit_krw_candidates = [normalize(x) for x in ['단가', 'unitkrw', 'unitcost', 'price', 'unit cost', 'unit cost (krw)', 'unit cost(krw)']]
        amount_krw_candidates = [normalize(x) for x in ['금액', 'amount', 'total', 'totalamount', 'amountkrw', 'total amount', 'total amount (krw)', 'total amount(krw)']]
        remark_candidates = [normalize(x) for x in ['비고', 'remark', 'note']]
        for row_idx in range(5, 31):
            colnames = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col).value
                if cell_value:
                    colnames.append(normalize(cell_value))
                else:
                    colnames.append('')
            # description, qty, unit, unit_krw, amount_krw, remark 컬럼 동적 매핑
            for col, name in enumerate(colnames, 1):
                if 'description' not in columns and name in desc_candidates:
                    columns['description'] = col
                if 'qty' not in columns and name in qty_candidates:
                    columns['qty'] = col
                if 'unit' not in columns and name in unit_candidates:
                    columns['unit'] = col
                if 'unit_krw' not in columns and name in unit_krw_candidates:
                    columns['unit_krw'] = col
                if 'amount_krw' not in columns and name in amount_krw_candidates:
                    columns['amount_krw'] = col
                if 'remark' not in columns and name in remark_candidates:
                    columns['remark'] = col
            # description, amount_krw, qty 등 주요 컬럼이 2개 이상 매핑되면 헤더로 간주
            match_count = sum(1 for k in ['description', 'qty', 'unit_krw', 'amount_krw'] if k in columns)
            if match_count >= 2:
                start_row = row_idx + 1
                break
        if not start_row or not columns:
            return categories
        SUMMARY_KEYWORDS = [
            '합계', '총액', 'vat', '참고', '소계', 'subtotal', 'total', 'sum', 'note', 'agency fee', 'cut', 'other'
        ]
        for row in range(start_row, ws.max_row + 1):
            row_data = {
                key: self.get_cell_value(ws, row, col)
                for key, col in columns.items()
            }
            desc = (str(row_data.get('description')) or '').strip()
            # summary 키워드가 포함된 행은 무조건 건너뜀
            if any(kw in desc.replace(' ', '').lower() for kw in SUMMARY_KEYWORDS):
                continue
            # 카테고리(섹션) 행: description만 있고 amount_krw만 있는 행
            if desc and not any(row_data.get(k) for k in ['unit_krw', 'qty', 'unit']) and row_data.get('amount_krw'):
                if current_category:
                    categories.append({
                        'category': current_category,
                        'category_total': self.clean_number(current_category_total),
                        'items': current_items
                    })
                current_category = desc
                current_category_total = row_data.get('amount_krw')
                current_items = []
                continue
            # 품목 데이터: description, 단가, 수량, 금액 등 주요 값이 있으면
            if desc and (
                row_data.get('unit_krw') or row_data.get('qty') or row_data.get('amount_krw')
            ):
                item = {
                    "description": desc,
                    "unit_krw": self.clean_number(row_data.get('unit_krw')),
                    "qty": self.clean_number(row_data.get('qty')),
                    "unit": row_data.get('unit'),
                    "amount_krw": self.clean_number(row_data.get('amount_krw')),
                    "remark": row_data.get('remark'),
                }
                current_items.append(item)
        if current_category:
            categories.append({
                'category': current_category,
                'category_total': self.clean_number(current_category_total),
                'items': current_items
            })
        return categories
    
    def extract_invoice(self, excel_path, template_type=None):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active
            if template_type is None:
                template_type_detected = self.detect_template_type(ws)
            else:
                template_type_detected = template_type
            template = self.templates.get(template_type_detected)
            if not template:
                print(f"알 수 없는 양식: {excel_path}")
                return None
            data = {
                'template_type': template_type_detected,
                'header': self.extract_header(ws, template),
                'items': self.extract_items(ws, template)
            }
            return data
        except Exception as e:
            print(f"견적서 추출 실패 ({excel_path}): {e}")
            return None

def main():
    # 분류 폴더별 템플릿 타입 매핑
    folder_template_map = {
        '영문INVOICE': '영문 INVOICE',
        'Quotation': 'Quotation',
        '한글견적서': '한글 견적서',
        '카테고리견적서': '카테고리 견적서',
        '거래명세서': '거래명세서',
    }
    base_dir = './분류결과'
    output_dir = os.path.join(base_dir, 'json')
    os.makedirs(output_dir, exist_ok=True)
    extractor = InvoiceExtractor()
    for folder, template_type in folder_template_map.items():
        folder_path = os.path.join(base_dir, folder)
        if not os.path.isdir(folder_path):
            continue
        for filename in os.listdir(folder_path):
            if not filename.endswith(('.xlsx', '.xls')):
                continue
            excel_path = os.path.join(folder_path, filename)
            print(f"[{template_type}] 처리 중: {filename}")
            data = extractor.extract_invoice(excel_path, template_type=template_type)
            if data:
                output_path = os.path.join(
                    output_dir,
                    f"{os.path.splitext(filename)[0]}.json"
                )
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"✅ 변환 완료: {output_path}")

if __name__ == "__main__":
    main() 