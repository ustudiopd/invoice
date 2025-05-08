import sys
import os
import json
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QTableWidget, QTableWidgetItem,
    QFrame, QTextBrowser, QTextEdit, QSplitter, QLabel,
    QMessageBox, QStyledItemDelegate
)
from PyQt5.QtGui import QColor, QFont, QPen
from PyQt5.QtCore import Qt
from dotenv import load_dotenv
import openpyxl
import requests


# .env에서 모든 환경변수 불러오기
load_dotenv()


# Dropbox 관련 환경변수
DROPBOX_APP_KEY = os.getenv("DROPBOX_APP_KEY", "")
DROPBOX_APP_SECRET = os.getenv("DROPBOX_APP_SECRET", "")
DROPBOX_ACCESS_TOKEN = os.getenv("DROPBOX_ACCESS_TOKEN", "")
DROPBOX_REFRESH_TOKEN = os.getenv("DROPBOX_REFRESH_TOKEN", "")
DROPBOX_SHARED_FOLDER_ID = os.getenv("DROPBOX_SHARED_FOLDER_ID", "")
DROPBOX_SHARED_FOLDER_NAME = os.getenv("DROPBOX_SHARED_FOLDER_NAME", "")
LOCAL_BID_FOLDER = os.getenv("LOCAL_BID_FOLDER", "")


# ChatGPT 관련 환경변수
GPT_API_KEY = os.getenv("CHATGPT_API_KEY", "")
GPT_MODEL = os.getenv("CHATGPT_MODEL", "gpt-4.1-mini")


def apply_tint(hex_rgb, tint):
    """hex_rgb: 'RRGGBB', tint: -1.0~1.0"""
    ch = [int(hex_rgb[0:2], 16), int(hex_rgb[2:4], 16), int(hex_rgb[4:6], 16)]
    out = [0, 0, 0]
    for i in range(3):
        c = ch[i]
        if tint < 0:
            nc = c * (1 + tint)
        else:
            nc = c * (1 - tint) + 255 * tint
        out[i] = max(0, min(int(round(nc)), 255))
    return QColor(out[0], out[1], out[2])


class BorderDelegate(QStyledItemDelegate):
    """openpyxl border 정보에 따라 셀 테두리만 그려주는 Delegate"""
    def paint(self, painter, option, index):
        # 1) 기본 렌더링 (배경·글자 등)
        super().paint(painter, option, index)
        # 2) border_info에 따라 테두리만 그림
        border_info = index.data(Qt.UserRole)
        if not isinstance(border_info, dict):
            return
        pen = QPen(QColor('#555555'))
        pen.setWidth(1)
        painter.save()
        painter.setPen(pen)
        r = option.rect
        # 위쪽 테두리
        if border_info.get("top"):
            painter.drawLine(r.topLeft(), r.topRight())
        # 아래쪽
        if border_info.get("bottom"):
            painter.drawLine(r.bottomLeft(), r.bottomRight())
        # 왼쪽
        if border_info.get("left"):
            painter.drawLine(r.topLeft(), r.bottomLeft())
        # 오른쪽
        if border_info.get("right"):
            painter.drawLine(r.topRight(), r.bottomRight())
        painter.restore()


class ZoomableTableWidget(QTableWidget):
    """확대/축소가 가능한 테이블 위젯"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.zoom_factor = 1.0
        self.min_zoom = 0.1  # 10%
        self.max_zoom = 2.5  # 250%
        self.zoom_step = 0.1
        self.base_font_size = 9
        self.base_row_height = 20
        self.last_mouse_pos = None
        # 특정 배율 목록 (90%, 100%, 125%, 150%, 175%, 200%, 250%)
        self.snap_zoom_levels = [0.9, 1.0, 1.25, 1.5, 1.75, 2.0, 2.5]
        self.snap_threshold = 0.05  # 스냅 감지 임계값

    def wheelEvent(self, event):
        """Ctrl + 휠로 확대/축소"""
        if event.modifiers() == Qt.ControlModifier:
            # 마우스 커서 위치 저장
            cursor_pos = event.pos()
            self.last_mouse_pos = cursor_pos
            
            # 현재 뷰포트에서의 상대적 위치 계산
            viewport_rect = self.viewport().rect()
            cursor_rel_x = cursor_pos.x() / viewport_rect.width()
            cursor_rel_y = cursor_pos.y() / viewport_rect.height()
            
            # 현재 스크롤 위치 저장
            h_scroll = self.horizontalScrollBar()
            v_scroll = self.verticalScrollBar()
            
            # 휠 델타값으로 확대/축소 방향 결정
            delta = event.angleDelta().y()
            
            # 휠 방향에 따라 확대/축소
            if delta > 0:  # 휠 업: 확대
                new_zoom = min(
                    self.zoom_factor + self.zoom_step,
                    self.max_zoom
                )
            elif delta < 0:  # 휠 다운: 축소
                new_zoom = max(
                    self.zoom_factor - self.zoom_step,
                    self.min_zoom
                )

            # 스냅 기능 적용
            new_zoom = self._apply_snap_zoom(new_zoom)
            
            # 확대/축소 비율이 변경된 경우에만 처리
            if new_zoom != self.zoom_factor:
                self.zoom_factor = new_zoom
                
                # 확대/축소 적용
                self._apply_zoom()
                
                # 새로운 스크롤 위치 계산
                new_h_max = h_scroll.maximum()
                new_v_max = v_scroll.maximum()
                
                # 마우스 커서 위치 기준으로 스크롤 위치 조정
                new_h_value = int(new_h_max * cursor_rel_x)
                new_v_value = int(new_v_max * cursor_rel_y)
                
                # 스크롤 위치 업데이트
                h_scroll.setValue(new_h_value)
                v_scroll.setValue(new_v_value)
            
            event.accept()
        else:
            # Ctrl 키가 눌려있지 않으면 기본 스크롤 동작
            super().wheelEvent(event)

    def _apply_snap_zoom(self, zoom_value):
        """가장 가까운 스냅 레벨로 확대/축소 값 조정"""
        closest_snap = None
        min_diff = float('inf')
        
        for snap_level in self.snap_zoom_levels:
            diff = abs(zoom_value - snap_level)
            if diff < min_diff:
                min_diff = diff
                closest_snap = snap_level
        
        if min_diff < self.snap_threshold:
            return closest_snap
        return zoom_value

    def set_zoom(self, value):
        """슬라이더로부터 확대/축소 값 설정"""
        # 현재 뷰포트의 중심점 계산
        viewport_rect = self.viewport().rect()
        center_x = viewport_rect.width() / 2
        center_y = viewport_rect.height() / 2
        
        # 현재 스크롤 위치 저장
        h_scroll = self.horizontalScrollBar()
        v_scroll = self.verticalScrollBar()
        
        # 새로운 확대/축소 값 계산
        zoom_value = value / 100.0
        new_zoom = self._apply_snap_zoom(zoom_value)
        
        # 확대/축소 비율이 변경된 경우에만 처리
        if new_zoom != self.zoom_factor:
            self.zoom_factor = new_zoom
            
            # 확대/축소 적용
            self._apply_zoom()
            
            # 새로운 스크롤 위치 계산
            new_h_max = h_scroll.maximum()
            new_v_max = v_scroll.maximum()
            
            # 중심점 기준으로 스크롤 위치 조정
            new_h_value = int(new_h_max * (center_x / viewport_rect.width()))
            new_v_value = int(new_v_max * (center_y / viewport_rect.height()))
            
            # 스크롤 위치 업데이트
            h_scroll.setValue(new_h_value)
            v_scroll.setValue(new_v_value)

    def _apply_zoom(self):
        """현재 zoom_factor를 적용"""
        # 폰트 크기 조정
        font = self.font()
        font.setPointSizeF(self.base_font_size * self.zoom_factor)
        self.setFont(font)
        
        # 행 높이와 열 너비를 동시에 조정
        for row in range(self.rowCount()):
            self.setRowHeight(
                row,
                int(self.base_row_height * self.zoom_factor)
            )
        
        for col in range(self.columnCount()):
            current_width = self.columnWidth(col)
            self.setColumnWidth(
                col,
                int(current_width * self.zoom_factor)
            )


class ExcelGPTViewer(QMainWindow):
    def log(self, msg):
        if hasattr(self, 'log_output'):
            self.log_output.append(msg)
        print(msg)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel + GPT 분석기")
        self.setGeometry(100, 100, 1400, 900)
        self.json_path = None
        self.excel_path = None

        # 메인 레이아웃
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # QSplitter(좌:엑셀, 우:챗)
        self.splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(self.splitter)

        # 1) 엑셀 뷰어 패널
        excel_panel = QWidget()
        excel_layout = QVBoxLayout(excel_panel)
        excel_layout.setContentsMargins(8, 8, 8, 8)
        
        # 상단 버튼 레이아웃
        top_layout = QHBoxLayout()
        file_btn = QPushButton("엑셀 파일 열기")
        file_btn.clicked.connect(self.open_excel)
        top_layout.addWidget(file_btn)
        excel_layout.addLayout(top_layout)
        
        # ZoomableTableWidget으로 변경
        self.excel_view = ZoomableTableWidget()
        # 기본 그리드(격자) 끄기
        self.excel_view.setShowGrid(False)
        # 셀별 테두리 그리도록 Delegate 설정
        self.excel_view.setItemDelegate(BorderDelegate(self.excel_view))
        self.excel_view.cellChanged.connect(self.on_cell_changed)
        excel_layout.addWidget(self.excel_view)
        
        self.splitter.addWidget(excel_panel)

        # 2) 챗봇 패널
        chat_frame = QFrame()
        chat_layout = QVBoxLayout(chat_frame)
        chat_layout.setContentsMargins(8, 8, 8, 8)
        chat_layout.setSpacing(8)
        self.model_label = QLabel(f"모델: {GPT_MODEL}")
        self.model_label.setStyleSheet("font-weight:bold;color:#0057b8;")
        chat_layout.addWidget(self.model_label)
        self.chat_output = QTextBrowser()
        self.chat_output.setOpenExternalLinks(True)
        self.chat_output.setMinimumHeight(180)
        chat_layout.addWidget(self.chat_output)
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        chat_layout.addWidget(line)
        self.chat_input = QTextEdit()
        self.chat_input.setPlaceholderText("질문을 입력하세요...")
        self.chat_input.setMinimumHeight(40)
        self.chat_input.setStyleSheet("border:2px solid #000;")
        chat_layout.addWidget(self.chat_input)
        btn_layout = QHBoxLayout()
        btn_layout.addStretch(1)
        self.send_btn = QPushButton("질문하기")
        self.send_btn.clicked.connect(self.ask_gpt)
        btn_layout.addWidget(self.send_btn)
        btn_layout.addStretch(1)
        btn_widget = QWidget()
        btn_widget.setLayout(btn_layout)
        chat_layout.addWidget(btn_widget)
        self.splitter.addWidget(chat_frame)
        self.splitter.setSizes([900, 500])

        # 로그 메시지창 (하단)
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setMaximumHeight(80)
        style = "background:#222;color:#eee;font-size:12px;"
        self.log_output.setStyleSheet(style)
        # 메인 레이아웃에 로그창 추가 (세로로 쌓기)
        vbox = QVBoxLayout()
        vbox.setContentsMargins(0, 0, 0, 0)
        vbox.setSpacing(0)
        vbox.addLayout(main_layout)
        vbox.addWidget(self.log_output)
        main_widget.setLayout(vbox)

        # 헤더/그리드 스타일
        self.excel_view.setStyleSheet("""
        QHeaderView::section {
            background-color: #3F4A73;
            color: white;
            font-weight: bold;
            border: 1px solid #555;
        }
        QTableWidget {
            gridline-color: #AAAAAA;
        }
        """)

    def open_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "엑셀 파일 선택", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        self.log(f"[작업] 엑셀 파일 열기: {path}")
        try:
            self.excel_path = path
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            # accent1 파랑 계열 RGB를 하드코딩
            accent_colors = {1: '3B4E87'}
            rows, cols = ws.max_row, ws.max_column
            table = self.excel_view
            table.blockSignals(True)
            table.clear()
            table.setRowCount(rows)
            table.setColumnCount(cols)

            # 헤더 키워드 정의 (부분 일치, 한글/영문 혼용)
            header_keywords = [
                "item", "품목", "항목", "품명",
                "상세 내역", "상 세 내 역", "상세내역",
                "description"
            ]
            quantity_keywords = ["quantity", "수량", "quant", "qty"]
            day_keywords = ["day", "일수"]
            unit_cost_keywords = [
                "unit cost", "단가", "unit krw", "unit"
            ]
            total_amount_keywords = [
                "total amount", "금액", "합계", "amount"
            ]

            header_row = None
            header_map = {}
            for r in range(1, ws.max_row + 1):
                row_values = [
                    str(ws.cell(row=r, column=c).value).strip().lower()
                    if ws.cell(row=r, column=c).value is not None else ""
                    for c in range(1, ws.max_column + 1)
                ]
                self.log(f"row {r} values: {row_values}")
                for idx, val in enumerate(row_values):
                    if any(k in val for k in header_keywords):
                        # '상세 내역'이 있으면 Description, 아니면 Item
                        if "상세" in val:
                            header_map["Description"] = idx
                        else:
                            header_map["Item"] = idx
                    if any(k in val for k in quantity_keywords):
                        header_map["Quantity"] = idx
                    if any(k in val for k in day_keywords):
                        header_map["day"] = idx
                    if any(k in val for k in unit_cost_keywords):
                        header_map["Unit Cost"] = idx
                    if any(k in val for k in total_amount_keywords):
                        header_map["Total Amount"] = idx
                if (
                    ("Item" in header_map or "Description" in header_map) and
                    "Quantity" in header_map and
                    "day" in header_map and
                    "Unit Cost" in header_map and
                    "Total Amount" in header_map
                ):
                    header_row = r
                    break

            items = []
            if header_row:
                for r in range(header_row + 1, ws.max_row + 1):
                    # 한 행의 모든 셀 텍스트를 합쳐 description으로 사용
                    row_texts = [
                        str(ws.cell(row=r, column=c+1).value).strip()
                        for c in range(ws.max_column)
                        if (
                            ws.cell(row=r, column=c+1).value is not None and
                            str(ws.cell(row=r, column=c+1).value).strip() != ""
                        )
                    ]
                    item_name = " | ".join(row_texts) if row_texts else None
                    # 불필요한 행(합계, 참고, VAT, 총액 등) 및 빈/짧은 행 제외
                    if (
                        not item_name or
                        str(item_name).strip() == "" or
                        len(str(item_name).strip()) <= 2 or
                        any(x in str(item_name) for x in [
                            "합계", "총액", "vat", "참고"
                        ])
                    ):
                        continue
                    quantity = ws.cell(
                        row=r, column=header_map.get("Quantity") + 1
                    ).value if header_map.get("Quantity") is not None else None
                    day = ws.cell(
                        row=r, column=header_map.get("day") + 1
                    ).value if header_map.get("day") is not None else None
                    unit_cost = ws.cell(
                        row=r,
                        column=header_map.get("Unit Cost") + 1
                    ).value if header_map.get("Unit Cost") is not None else None
                    total_amount = ws.cell(
                        row=r,
                        column=header_map.get("Total Amount") + 1
                    ).value if header_map.get("Total Amount") is not None else None

                    item = {
                        "description": str(item_name),
                        "quantity": quantity,
                        "day": day,
                        "unit_price": unit_cost,
                        "amount": total_amount
                    }
                    items.append(item)

            # 셀 값, 폰트, 배경, 정렬, 테두리(UserRole에 정보 저장)
            for r in range(1, rows+1):
                for c in range(1, cols+1):
                    cell = ws.cell(r, c)
                    text = cell.value or ""
                    item = QTableWidgetItem(str(text))

                    # 폰트
                    f = cell.font
                    point_size = int(f.sz) if f.sz is not None else -1
                    qf = QFont(f.name, point_size)
                    qf.setBold(f.b)
                    qf.setItalic(f.i)
                    item.setFont(qf)

                    # 배경색 (Excel 테마 컬러 + 틴트 지원)
                    color = None
                    fill = cell.fill
                    fg = fill.fgColor if hasattr(fill, 'fgColor') else None
                    try:
                        if (fill.patternType in
                                ('solid', 'gray125', 'darkGrid', 'lightGrid')
                                and fg):
                            # 1) Theme 컬러(무조건 파랑 accent1) + 틴트
                            if fg.type == 'theme':
                                tint = getattr(fg, 'tint', 0.0)
                                hex_rgb = accent_colors[1]
                                color = apply_tint(hex_rgb, tint)
                                msg = (
                                    f"[THEME+TINT:파랑] ({r},{c}) "
                                    f"tint={tint} {hex_rgb} → "
                                    f"({color.red()},{color.green()},"
                                    f"{color.blue()}) 적용"
                                )
                                self.log(msg)
                            # 2) RGB 컬러
                            elif fg.type == 'rgb' and fg.rgb:
                                rgb = (
                                    fg.rgb[2:]
                                    if fg.rgb.startswith('FF')
                                    else fg.rgb
                                )
                                color = QColor(
                                    int(rgb[0:2], 16),
                                    int(rgb[2:4], 16),
                                    int(rgb[4:6], 16)
                                )
                                msg = (
                                    f"[RGB] ({r},{c}) {rgb} → "
                                    f"({color.red()},{color.green()},"
                                    f"{color.blue()}) 적용"
                                )
                                self.log(msg)
                            # 3) Indexed 컬러
                            elif (fg.type == 'indexed' and
                                fg.indexed is not None):
                                from openpyxl.styles.colors import COLOR_INDEX
                                idx = fg.indexed
                                if 0 <= idx < len(COLOR_INDEX):
                                    hexcol = COLOR_INDEX[idx][2:]
                                    color = QColor(
                                        int(hexcol[0:2], 16),
                                        int(hexcol[2:4], 16),
                                        int(hexcol[4:6], 16)
                                    )
                                    msg = (
                                        f"[INDEXED] ({r},{c}) idx={idx} "
                                        f"{hexcol} → ({color.red()},"
                                        f"{color.green()},{color.blue()}) 적용"
                                    )
                                    self.log(msg)
                            # 4) Gradient Fill (첫 stop만 사용)
                            elif (hasattr(fill, 'gradientType')
                                and fill.gradientType):
                                stops = getattr(fill, 'stop', None)
                                if (stops and hasattr(stops[0], 'color')
                                        and hasattr(stops[0].color, 'rgb')):
                                    rgb = (
                                        stops[0].color.rgb[2:]
                                        if stops[0].color.rgb.startswith('FF')
                                        else stops[0].color.rgb
                                    )
                                    color = QColor(
                                        int(rgb[0:2], 16),
                                        int(rgb[2:4], 16),
                                        int(rgb[4:6], 16)
                                    )
                                    msg = (
                                        f"[GRADIENT] ({r},{c}) {rgb} → "
                                        f"({color.red()},{color.green()},"
                                        f"{color.blue()}) 적용"
                                    )
                                    self.log(msg)
                    except Exception as e:
                        self.log(f"[ERROR] ({r},{c}) 색상 파싱 오류: {e}")
                    if (color and (color.red(), color.green(), color.blue())
                            != (255, 255, 255)):
                        item.setBackground(color)

                    # 정렬
                    align = cell.alignment
                    qt_align = 0
                    if align.horizontal == 'center':
                        qt_align |= Qt.AlignHCenter
                    elif align.horizontal == 'right':
                        qt_align |= Qt.AlignRight
                    else:
                        qt_align |= Qt.AlignLeft
                    if align.vertical == 'center':
                        qt_align |= Qt.AlignVCenter
                    elif align.vertical == 'bottom':
                        qt_align |= Qt.AlignBottom
                    else:
                        qt_align |= Qt.AlignTop
                    item.setTextAlignment(qt_align)

                    # openpyxl Border 객체를 dict로 변환해 UserRole에 저장
                    b = cell.border
                    border_info = {
                        "top": bool(b.top and b.top.style),
                        "bottom": bool(b.bottom and b.bottom.style),
                        "left": bool(b.left and b.left.style),
                        "right": bool(b.right and b.right.style),
                    }
                    item.setData(Qt.UserRole, border_info)
                    table.setItem(r-1, c-1, item)

            # 병합 셀
            for merged in ws.merged_cells.ranges:
                r0, c0 = merged.min_row-1, merged.min_col-1
                rs = merged.max_row - merged.min_row + 1
                cs = merged.max_col - merged.min_col + 1
                table.setSpan(r0, c0, rs, cs)

            # 열 너비
            for idx, col_dim in ws.column_dimensions.items():
                col = openpyxl.utils.column_index_from_string(idx) - 1
                if col < cols and col_dim.width:
                    table.setColumnWidth(col, int(col_dim.width * 7))
            # 행 높이
            for r, row_dim in ws.row_dimensions.items():
                if row_dim.height is not None and r-1 < rows:
                    table.setRowHeight(r-1, int(row_dim.height * 1.2))

            # 금액 열 읽기전용(예시: 5번째 열)
            amount_col = 4
            for r in range(rows):
                it = table.item(r, amount_col)
                if it:
                    it.setFlags(it.flags() & ~Qt.ItemIsEditable)

            table.blockSignals(False)

            # JSON 파일 경로
            json_path = os.path.splitext(path)[0] + ".json"
            self.json_path = json_path
            
            # 1) 기존 테이블 기반 JSON 변환
            data_dict = self._widget_to_json_schema()
            
            # 2) meta 필드에 파일명 저장
            data_dict.setdefault("meta", {})
            data_dict["meta"]["file_name"] = os.path.basename(path)
            
            # 3) 헤더 정보: 워크시트 상단에서 라벨-값 쌍 자동 추출
            hdr = {}
            # 지원할 라벨 키워드 목록 (기존/새 포맷 혼용)
            header_labels = [
                "DATE", "QUOTATION #", "Payment date", "SHIP TO",
                "발급일", "공급자", "등록번호", "상호", "대표이사", "사업자 주소"
            ]
            # 상단 1~15행, 전체 열을 뒤져서, 키가 발견되면 오른쪽 셀 또는 아래 셀 값을 가져옴
            for r in range(1, 16):
                for c in range(1, ws.max_column):
                    raw = ws.cell(row=r, column=c).value
                    if raw is None:
                        continue
                    key = str(raw).strip().rstrip(":")
                    if key in header_labels:
                        # 오른쪽 셀 우선, 없으면 바로 아래 셀
                        val = ws.cell(row=r, column=c+1).value \
                            or ws.cell(row=r+1, column=c).value
                        hdr[key] = val
            data_dict["meta"]["header"] = hdr
            
            # 4) 요약 정보: 워크시트 전체에서 합계/세금 등 라벨-값 자동 추출
            summary = {}
            # 내부 필드명: [매칭할 키워드 리스트]
            summary_labels = {
                "subtotal":   ["Sub total", "소계"],
                "tax_rate":   ["Tax rate", "세율"],
                "tax_due":    ["Tax due", "세금", "Tax due"],
                "other":      ["Other", "기타"],
                "total_due":  ["TOTAL Due", "총액", "합계", "TOTAL"]
            }
            # 전체 영역 순회
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column):
                    raw = ws.cell(row=r, column=c).value
                    if raw is None:
                        continue
                    txt = str(raw).strip()
                    for fld, keys in summary_labels.items():
                        if any(k in txt for k in keys):
                            # 값은 오른쪽 셀에서 가져오기
                            summary[fld] = ws.cell(row=r, column=c+1).value
            data_dict["summary"] = summary

            # 5) 상단 고정 레이블: D3:E9 범위 읽어 header로 저장
            header = {}
            for r in range(3, 10):  # 행 3부터 9까지
                raw_key = ws.cell(row=r, column=4).value  # D열
                if raw_key:
                    key = str(raw_key).strip().rstrip(':')
                    header[key] = ws.cell(row=r, column=5).value  # E열
            data_dict["header"] = header
            
            data_dict["items"] = items  # 추출된 품목을 반드시 저장
            self.log(
                "최종 저장될 JSON 데이터: "
                + json.dumps(
                    data_dict,
                    ensure_ascii=False,
                    indent=2,
                    default=str
                )
            )
            # 6) JSON 파일 저장
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(
                    data_dict,
                    f,
                    ensure_ascii=False,
                    indent=2,
                    default=str
                )
            self.log(f"[작업] JSON 파일 저장: {json_path}")

            data_dict["items"] = items  # 추출된 품목을 명확히 할당
            self.log(f"추출된 품목: {items}")
            self.log(f"헤더 행: {header_row}, header_map: {header_map}")
        except Exception as e:
            self.log(f"[오류] 엑셀 파일 열기 실패: {e}")
            QMessageBox.critical(
                self,
                "엑셀 파일 오류",
                f"엑셀 파일을 불러올 수 없습니다:\n{e}"
            )

    def on_cell_changed(self, row, col):
        if not self.json_path:
            return
        data_dict = self._widget_to_json_schema()
        with open(self.json_path, "w", encoding="utf-8") as f:
            json.dump(data_dict, f, ensure_ascii=False, indent=2)
        self.log(f"[작업] 셀 변경: ({row},{col}) → JSON 동기화")

    def _widget_to_json_schema(self):
        table = self.excel_view
        result = {
            "meta": {},
            "items": [],
            "discounts": [],
            "summary": {},
            "comments": ""
        }
        current_category = None
        for row in range(table.rowCount()):
            a_item = table.item(row, 0)
            d_item = table.item(row, 3)
            a = a_item.text() if a_item else ""
            d = d_item.text() if d_item else ""
            # 1) 섹션 헤더
            if a and not d:
                current_category = a.strip()
                continue
            # 2) 품목 행
            if a and d:
                try:
                    item = {
                        "category": current_category,
                        "description": a.strip(),
                        "unit_price": (
                            float(table.item(row, 1).text())
                            if table.item(row, 1)
                            and table.item(row, 1).text()
                            else 0
                        ),
                        "quantity": (
                            float(table.item(row, 2).text())
                            if table.item(row, 2)
                            and table.item(row, 2).text()
                            else 0
                        ),
                        "unit_count": float(table.item(row, 3).text()),
                        "amount": (
                            float(table.item(row, 4).text())
                            if table.columnCount() > 4
                            and table.item(row, 4)
                            and table.item(row, 4).text()
                            else None
                        )
                    }
                    result["items"].append(item)
                except Exception:
                    continue
                continue
            # 3) 요약/할인/합계 행 (A열 비어있고 D열에 숫자)
            if not a and d:
                try:
                    val = float(d)
                    if val < 0:
                        result["discounts"].append({
                            "description": "",
                            "amount": -val
                        })
                    else:
                        if "subtotal" not in result["summary"]:
                            result["summary"]["subtotal"] = val
                        elif "tax_amount" not in result["summary"]:
                            result["summary"]["tax_amount"] = val
                        else:
                            result["summary"]["total_due"] = val
                except Exception:
                    continue
                continue
        return result

    def ask_gpt(self):
        user_q = self.chat_input.toPlainText().strip()
        if not user_q or not self.json_path:
            return
        with open(self.json_path, "r", encoding="utf-8") as f:
            quotation_json = f.read()
        messages = [
            {
                "role": "system",
                "content": "아래 견적서 JSON을 참고해 질문에 답변해 주세요."
            },
            {
                "role": "user",
                "content": (
                    f"견적서 데이터:\n```json\n{quotation_json}\n```\n"
                    f"질문: {user_q}"
                )
            }
        ]
        answer = ask_gpt_api(messages, GPT_API_KEY, GPT_MODEL)
        self.chat_output.append(f"<b>질문:</b> {user_q}")
        self.chat_output.append("<b>GPT:</b>")
        self.chat_output.append(answer)
        self.chat_input.clear()
        self.log("[작업] GPT 질문 전송 및 응답 수신 완료")


def ask_gpt_api(messages, api_key, model):
    """GPT API를 호출하여 응답을 받아옵니다."""
    if not api_key:
        return "[OpenAI API 키를 .env에 입력하세요]"
    url = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    data = {
        "model": model,
        "messages": messages,
        "max_tokens": 2048,
        "temperature": 0.7
    }
    try:
        resp = requests.post(url, headers=headers, json=data, timeout=30)
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        return f"[GPT 호출 오류] {e}"


if __name__ == "__main__":
    app = QApplication(sys.argv)
    viewer = ExcelGPTViewer()
    viewer.show()
    sys.exit(app.exec_())